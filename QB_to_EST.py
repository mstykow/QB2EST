#! python3
# Program converts QuickBooks contacts export Excel file into a text file that
# can be imported with Canada Posts's EST 2.0 software.

import openpyxl, re, warnings, pycountry, provinces, os
from openpyxl.cell import get_column_letter, column_index_from_string

def none_to_string(cellValue):
    if cellValue == None:
        return ''
    else:
        return cellValue

addressRegex = re.compile(r'''(
    ([^,]+)           # city (everything up to first comma)
    ([, ]{2})         # separator
    ([A-Z]{2}|[^,]+)  # province (either two letter code of name)
    (\s|, )?          # separator
    (.*)              # postal or zip code (everything up to the end)
    )''', re.VERBOSE)

print('''Please make sure that the QuickBooks export file is located in the same directory as this program.''')
print("Please enter the export file's filename (e.g. export.xlsx):")
exportFile = input()
    
warnings.simplefilter("ignore")
contactsWB = openpyxl.load_workbook(exportFile)
contactsSheet = contactsWB.get_sheet_by_name('Sheet1')
maxRow = contactsSheet.max_row
maxColumn = contactsSheet.max_column

# Make sheet with required fields.
reqFields = {
    'Record Type': 'A',
    'Client ID': 'B',
    'Title Name': 'C',
    'First Name': 'D',
    'Last Name': 'E',
    'Title/ Dept.': 'F',
    'Company Name': 'G',
    'Additional Address Information': 'H',
    'Address Line 1': 'I',
    'Address Line 2': 'J',
    'City': 'K',
    'Province or State': 'L',
    'Postal Code or Zip Code': 'M',
    'Country Code': 'N',
    'Client Voice Phone': 'O',
    'Client Fax Number': 'P',
    'Client Email Address': 'Q',
    'Tax ID / IRS / VAT': 'R',
    'Email # 2': 'S'
}

postSheet = contactsWB.create_sheet(title = 'EST')
for i in range(len(reqFields)):
    postSheet.cell(row = 1, column = i + 1).value = list(reqFields.keys())[i]

# Find top-left corner of data table.
startColumn = 2
startRow = 1
for cell in contactsSheet.columns[1]:
    if cell.value == '':
        startRow += 1
    else:
        break

# Find columns of source fields.
srcFieldsList = [
    'Ship to 1',
    'Ship to 2',
    'Ship to 3',
    'Ship to 4',
    'Ship to 5',
    'Main Phone',
    'Main Email'
    ]

srcFieldsDict = {}
for cell in contactsSheet.rows[startRow - 1]:
    if cell.value in srcFieldsList:
        srcFieldsDict[cell.value] = cell.column
    else:
        continue

# Bring "City, Province, Postal Code" data into the same column.
for i in range(startRow + 1, maxRow + 1):
    if contactsSheet[srcFieldsDict['Ship to 5'] + str(i)].value == None:
        if contactsSheet[srcFieldsDict['Ship to 4'] + str(i)].value == None:
            if contactsSheet[srcFieldsDict['Ship to 3'] + str(i)].value == None:
                continue
            else:
                contactsSheet[srcFieldsDict['Ship to 5'] + str(i)].value = contactsSheet[srcFieldsDict['Ship to 3'] + str(i)].value
                contactsSheet[srcFieldsDict['Ship to 4'] + str(i)].value = contactsSheet[srcFieldsDict['Ship to 2'] + str(i)].value
                contactsSheet[srcFieldsDict['Ship to 3'] + str(i)].value = None
                contactsSheet[srcFieldsDict['Ship to 2'] + str(i)].value = None
        else:
            contactsSheet[srcFieldsDict['Ship to 5'] + str(i)].value = contactsSheet[srcFieldsDict['Ship to 4'] + str(i)].value
            contactsSheet[srcFieldsDict['Ship to 4'] + str(i)].value = contactsSheet[srcFieldsDict['Ship to 3'] + str(i)].value
            contactsSheet[srcFieldsDict['Ship to 3'] + str(i)].value = None
    else:
        continue

# Convert countries into country codes.
for country in contactsSheet.columns[column_index_from_string(srcFieldsDict['Ship to 5']) - 1][startRow:]:
    try:
        countryInfo = pycountry.countries.lookup(country.value)
    except:
        countryInfo = None
    if countryInfo:
        country.value = countryInfo.alpha_2

# Map source fields to required fields and write to postSheet.
translator = {
    'Ship to 1': 'Last Name',
    'Ship to 2': 'Address Line 1',
    'Ship to 3': 'Address Line 2',
    'Ship to 4': 'City',
    'Ship to 5': 'Country Code',
    'Main Phone': 'Client Voice Phone',
    'Main Email': 'Client Email Address'
    }

for i in range(2, maxRow - (startRow - 1) + 1):
    for key in list(translator.keys()):
        postSheet[reqFields[translator[key]] + str(i)].value = contactsSheet[srcFieldsDict[key] + str(startRow + i - 1)].value

# Split "City, Province, Postal Code" information and write to postSheet.
for address in postSheet.columns[column_index_from_string(reqFields['City']) - 1][1:]:
    mo = addressRegex.search(str(address.value))
    if mo:
        postSheet[reqFields['Province or State'] + str(address.row)].value = mo.groups()[3]
        postSheet[reqFields['Postal Code or Zip Code'] + str(address.row)].value = mo.groups()[5]
        address.value = mo.groups()[1]
    else:
        continue

# Convert provinces into province codes.
for province in postSheet.columns[column_index_from_string(reqFields['Province or State']) - 1][1:]:
    if province.value in list(provinces.codes.keys()):
        province.value = provinces.codes[province.value]

# Select non-empty record types.
for name in postSheet.columns[column_index_from_string(reqFields['Last Name']) - 1][1:]:
    if name.value:
        postSheet[reqFields['Record Type'] + str(name.row)].value = '1'

# Write EST sheet data to .txt file with segments enclosed by quotes and separated by commas.
export = open('contacts.txt', 'w')
export.close()
for rowOfCells in postSheet.rows[1:]:
    rowOfValues = []
    for cell in rowOfCells:
        rowOfValues.append('"' + none_to_string(cell.value) + '"')
    export = open('contacts.txt', 'a')
    export.write(','.join(rowOfValues) + '\n')
    export.close()

contactsWB.save(exportFile)
