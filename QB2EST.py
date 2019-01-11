#! python3
# Program converts QuickBooks contacts export Excel file into a text file that
# can be imported with Canada Posts's EST 2.0 software.

# Install these modules before first time running script.
import openpyxl, pycountry
from openpyxl.cell import get_column_letter, column_index_from_string

# Python standard modules.
import re, os, warnings
warnings.simplefilter("ignore")

# External module part of this program.
import provinces

# Opens sheet with the data, determines the last row, and creates import filename.
def export_file_processor(exportFile):
    while os.path.isfile(exportFile) == False:
        print("Filename does not exist. Please try again.")
        exportFile = input()    
    print("Processing...")
    workbook = openpyxl.load_workbook(exportFile)
    worksheet = workbook.get_sheet_by_name('Sheet1')
    maxRow = worksheet.max_row
    filename = os.path.splitext(exportFile)[0] + '.txt'
    return workbook, worksheet, maxRow, filename

# Creates new sheet with a given list of fields inside a workbook.
def make_import_sheet(workbook, name, fieldList):
    worksheet = workbook.create_sheet(title = name)
    for i in range(len(fieldList)):
        worksheet.cell(row = 1, column = i + 1).value = list(fieldList.keys())[i]
    return worksheet

# Finds the first row containing data in column 'B' of a sheet.
def find_data(sheet):
    startRow = 1
    for cell in sheet.columns[1]:
        if cell.value == None:
            startRow += 1
        else:
            break
    return startRow

# Finds columns corresponding to each cell in a list and returns assignment as a dictionary.
def find_columns(listOfCells, sheet):
    dictionary = {}
    for cell in sheet.rows[startRow - 1]:
        if cell.value in listOfCells:
            dictionary[cell.value] = cell.column
        else:
            continue
    return dictionary

# Right-aligns four columns of a spreadsheet.
def address_align(sheet, start, end, A, B, C, D):
    for i in range(start + 1, end + 1):
        if sheet[D + str(i)].value == None:
            if sheet[C + str(i)].value == None:
                if sheet[B + str(i)].value == None:
                    continue
                else:
                    sheet[D + str(i)].value = sheet[B + str(i)].value
                    sheet[C + str(i)].value = sheet[A + str(i)].value
                    sheet[B + str(i)].value = None
                    sheet[A + str(i)].value = None
            else:
                sheet[D + str(i)].value = sheet[C + str(i)].value
                sheet[C + str(i)].value = sheet[B + str(i)].value
                sheet[B + str(i)].value = None
        else:
            continue

# Converts country names or abbreviation in column of sheet to ISO 3166-2 standard code.
def country_converter(sheet, column, start):
    for country in sheet.columns[column_index_from_string(column) - 1][start:]:
        try:
            countryInfo = pycountry.countries.lookup(country.value)
        except:
            countryInfo = None
        if countryInfo:
            country.value = countryInfo.alpha_2

# Writes source cells in sheet1 to target cells in sheet2 via a mapping.
def source_to_target(sheet1, start, end, sheet2, mapping, sourceDict, targetDict):
    for i in range(2, end - (start - 1) + 1):
        for key in list(mapping.keys()):
            sheet2[targetDict[mapping[key]] + str(i)].value = sheet1[sourceDict[key] + \
                                                                     str(start + i - 1)].value

# Splits an address made of city, province, postal code into three pieces and writes these
# pieces to columns A, B, C.
def address_splitter(sheet, A, B, C):
    for address in sheet.columns[column_index_from_string(A) - 1][1:]:
        mo = addressRegex.search(str(address.value))
        if mo:
            sheet[B + str(address.row)].value = mo.groups()[3]
            sheet[C + str(address.row)].value = mo.groups()[5]
            address.value = mo.groups()[1]
        else:
            continue

# Converts province names in column of sheet to Canada Post standard code.
def province_converter(sheet, column):
    for province in sheet.columns[column_index_from_string(column) - 1][1:]:
        if province.value in list(provinces.codes.keys()):
            province.value = provinces.codes[province.value]

# Assigns column B of a row the value '1' if column A is non-empty.
def non_empty_names(sheet, A, B):
    for name in sheet.columns[column_index_from_string(A) - 1][1:]:
        if name.value:
            sheet[B + str(name.row)].value = '1'

# Turns a None valued cell into an empty cell.
def none_to_string(cellValue):
    if cellValue == None:
        return ''
    else:
        return cellValue

# Writes data from rows with nonempty first cell into filename where each value
# is separated by commas enclosed by quotes.
def quote_comma_export(filename, sheet):
    export = open(filename, 'w')
    export.close()
    for rowOfCells in sheet.rows[1:]:
        if rowOfCells[0].value == None:
            continue
        else:
            rowOfValues = []
            for cell in rowOfCells:
                rowOfValues.append('"' + none_to_string(cell.value) + '"')
            export = open(filename, 'a')
            export.write(','.join(rowOfValues) + '\n')
            export.close()

# Regex for splitting up City, Province, Postal Code
addressRegex = re.compile(r'''(
    ([^,]+)           # city (everything up to first comma)
    ([, ]{2})         # separator
    ([A-Z]{2}|[^,]+)  # province (either two letter code of name)
    (\s|, )?          # separator
    (.*)              # postal or zip code (everything up to the end)
    )''', re.VERBOSE)

# Fields required by EST 2.0 and their spreadsheet column assignment.
reqFields = {
    'Record Type': 'A',
    'Client ID': 'B',
    'Title Name': 'C',
    'First Name': 'D',
    'Last Name': 'E',
    'Title/ Dept.': 'F',
    'Company Name': 'G',
    'Additional Address Information': 'H',
    'Address Line 1': 'I',
    'Address Line 2': 'J',
    'City': 'K',
    'Province or State': 'L',
    'Postal Code or Zip Code': 'M',
    'Country Code': 'N',
    'Client Voice Phone': 'O',
    'Client Fax Number': 'P',
    'Client Email Address': 'Q',
    'Tax ID / IRS / VAT': 'R',
    'Email # 2': 'S'
}

# List of QuickBooks fields needed to populate EST sheet.
srcFieldsList = [
    'Ship to 1',
    'Ship to 2',
    'Ship to 3',
    'Ship to 4',
    'Ship to 5',
    'Main Phone',
    'Main Email'
    ]

# Mapping from QuickBooks export fields to EST required fields.
translator = {
    'Ship to 1': 'Last Name',
    'Ship to 2': 'Address Line 1',
    'Ship to 3': 'Address Line 2',
    'Ship to 4': 'City',
    'Ship to 5': 'Country Code',
    'Main Phone': 'Client Voice Phone',
    'Main Email': 'Client Email Address'
    }

# Get export file.
print('''Please make sure that the QuickBooks export file is located in the
same directory as this program.''')
print("Please enter the export file's filename (e.g. export.xlsx):")
exportFile = input()
contactsWB, contactsSheet, maxRow, importFile = export_file_processor(exportFile)

# Make sheet with required fields.
postSheet = make_import_sheet(contactsWB, 'EST', reqFields)

# Find top-left corner of data table.
startRow = find_data(contactsSheet)

# Find columns of source fields.
srcFieldsDict = find_columns(srcFieldsList, contactsSheet)

# Bring "City, Province, Postal Code" data into the same column.
srcCol1 = srcFieldsDict['Ship to 2']
srcCol2 = srcFieldsDict['Ship to 3']
srcCol3 = srcFieldsDict['Ship to 4']
srcCol4 = srcFieldsDict['Ship to 5']
address_align(contactsSheet, startRow, maxRow, srcCol1, srcCol2, srcCol3, srcCol4)

# Convert countries into country codes.
country_converter(contactsSheet, srcCol4, startRow)

# Map source fields to required fields and write to postSheet.
source_to_target(contactsSheet, startRow, maxRow, postSheet, translator, srcFieldsDict, reqFields)

# Split "City, Province, Postal Code" information and write to postSheet.
reqCol1 = reqFields['City']
reqCol2 = reqFields['Province or State']
reqCol3 = reqFields['Postal Code or Zip Code']
address_splitter(postSheet, reqCol1, reqCol2, reqCol3)

# Convert provinces into province codes.
province_converter(postSheet, reqCol2)

# Select non-empty record types.
reqCol4 = reqFields['Last Name']
reqCol5 = reqFields['Record Type']
non_empty_names(postSheet, reqCol4, reqCol5)

# Write EST sheet data to .txt file with segments enclosed by quotes and separated by commas.
quote_comma_export(importFile, postSheet)

print('EST import file has been created and is now available as %s.' % (importFile))

# Uncomment the following block to save import data to Excel file also.
# This may be useful for debugging the above code.

#try:
#    print('Saving Excel data...')
#    contactsWB.save(exportFile)
#    print('EST import data also saved to sheet "EST" of %s.' % (exportFile))
#except:
#    print('''Could not also save EST import data in %s
#          because it is open in another application.''' % (exportFile))
